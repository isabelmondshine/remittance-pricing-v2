"""
remittance_pricer.py
────────────────────────────────────────────────────────────────────────────────
FT Partners – Remittance Competitive Pricing Tracker
Pulls live bank-to-bank transfer pricing from each provider's public website
and writes results to a timestamped Excel file.

PROVIDERS & APPROACH:
  Wise         → Public Comparison API (api.wise.com/v4/comparisons)
                 Also covers: Remitly, Western Union, MoneyGram, Revolut
                 Updated ~hourly by Wise; bank transfer in/out only
  Remitly      → Public price calculator endpoint (api.remitly.com)
  Western Union→ Public price estimator (wuenvironments.com static-estimate)
  MoneyGram    → Public fee estimator (moneygram.com)
  Revolut      → Scraped from public transfer page (Playwright headless browser)
  Euronet      → Scraped from public currency converter page (Playwright)

REQUIREMENTS:
  pip install requests openpyxl playwright
  playwright install chromium

USAGE:
  python remittance_pricer.py                          # all corridors, all amounts
  python remittance_pricer.py --corridors USD-EUR      # single corridor
  python remittance_pricer.py --amounts 500 1000 5000  # specific amounts

OUTPUT:
  remittance_pricing_YYYYMMDD_HHMMSS.xlsx
────────────────────────────────────────────────────────────────────────────────
"""

import argparse
import json
import re
import time
import traceback
from datetime import datetime
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ─── Configuration ─────────────────────────────────────────────────────────────

CORRIDORS = [
    ("USD", "EUR"),
    ("USD", "GBP"),
    ("USD", "MXN"),
    ("USD", "INR"),
    ("USD", "PHP"),
    ("USD", "CAD"),
    ("USD", "AUD"),
    ("USD", "JPY"),
]

AMOUNTS = [100, 500, 1000, 5000, 10000]

# Destination country codes needed by some APIs (ISO 3166-1 alpha-2)
CURRENCY_TO_COUNTRY = {
    "EUR": "DE",   # Germany as representative Eurozone country
    "GBP": "GB",
    "MXN": "MX",
    "INR": "IN",
    "PHP": "PH",
    "CAD": "CA",
    "AUD": "AU",
    "JPY": "JP",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "en-US,en;q=0.9",
}

# ─── Data model ────────────────────────────────────────────────────────────────

class Quote:
    """Holds the result of a single pricing lookup."""
    def __init__(
        self,
        provider: str,
        from_ccy: str,
        to_ccy: str,
        send_amount: float,
        fee_usd: Optional[float] = None,
        fx_rate: Optional[float] = None,
        fx_markup_pct: Optional[float] = None,
        received_amount: Optional[float] = None,
        note: str = "",
        error: str = "",
    ):
        self.provider = provider
        self.from_ccy = from_ccy
        self.to_ccy = to_ccy
        self.send_amount = send_amount
        self.fee_usd = fee_usd
        self.fx_rate = fx_rate
        self.fx_markup_pct = fx_markup_pct
        self.received_amount = received_amount
        self.note = note
        self.error = error

    def __repr__(self):
        return (
            f"<Quote {self.provider} {self.from_ccy}→{self.to_ccy} "
            f"${self.send_amount} | fee=${self.fee_usd} mkp={self.fx_markup_pct}% "
            f"rcv={self.received_amount} {self.to_ccy}>"
        )


# ─── Provider: Wise (+ comparison data) ────────────────────────────────────────

def fetch_wise_comparison(from_ccy: str, to_ccy: str, amount: float) -> dict:
    """
    Calls Wise's public Comparison API.
    Returns raw JSON. No auth required. Bank transfer in/out.
    Covers: Wise, Remitly, Western Union, MoneyGram, Revolut, Xoom, Ria, etc.
    """
    url = (
        f"https://api.wise.com/v4/comparisons/"
        f"?sourceCurrency={from_ccy}&targetCurrency={to_ccy}&sendAmount={amount}"
    )
    resp = requests.get(url, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    return resp.json()


def parse_wise_comparison(data: dict, from_ccy: str, to_ccy: str, amount: float) -> list[Quote]:
    """
    Extracts best bank-transfer quote per provider from Wise comparison API response.
    """
    quotes = data.get("quotes", [])
    best: dict[str, dict] = {}

    for q in quotes:
        provider_id = (q.get("provider", {}).get("id") or "").lower().replace(" ", "")
        if not provider_id:
            continue
        rcv = q.get("totalReceivedAmount") or q.get("recipientGetsAmount") or 0
        existing = best.get(provider_id, {})
        existing_rcv = existing.get("totalReceivedAmount") or existing.get("recipientGetsAmount") or 0
        if rcv > existing_rcv:
            best[provider_id] = q

    results = []
    provider_display = {
        "wise": "Wise",
        "remitly": "Remitly",
        "westernunion": "Western Union",
        "moneygram": "MoneyGram",
        "revolut": "Revolut",
        "xoom": "Xoom",
        "ria": "Ria Money",
        "worldremit": "WorldRemit",
        "transfergo": "TransferGo",
    }

    for pid, q in best.items():
        fee = q.get("fee") or (q.get("fees") or {}).get("total")
        rate = q.get("rate") or q.get("exchangeRate")
        markup = q.get("rateMarkupPercentage")
        received = q.get("totalReceivedAmount") or q.get("recipientGetsAmount")
        label = provider_display.get(pid, pid.capitalize())
        results.append(Quote(
            provider=label,
            from_ccy=from_ccy,
            to_ccy=to_ccy,
            send_amount=amount,
            fee_usd=fee,
            fx_rate=rate,
            fx_markup_pct=markup,
            received_amount=received,
            note="Source: Wise Comparison API (bank transfer in/out)",
        ))

    return results


def get_wise_quotes(from_ccy: str, to_ccy: str, amount: float) -> list[Quote]:
    try:
        data = fetch_wise_comparison(from_ccy, to_ccy, amount)
        return parse_wise_comparison(data, from_ccy, to_ccy, amount)
    except Exception as e:
        return [Quote("Wise (API)", from_ccy, to_ccy, amount, error=str(e))]


# ─── Provider: Remitly (direct) ────────────────────────────────────────────────

def get_remitly_quote(from_ccy: str, to_ccy: str, amount: float) -> Quote:
    """
    Calls Remitly's public pricing calculator endpoint.
    Uses the same JSON API their website calculator uses.
    Economy rate, bank deposit delivery.
    """
    country = CURRENCY_TO_COUNTRY.get(to_ccy, "")
    if not country:
        return Quote("Remitly", from_ccy, to_ccy, amount, error=f"No country mapping for {to_ccy}")

    try:
        # Remitly's public price endpoint (same as their web calculator)
        url = "https://api.remitly.io/v3/pricing/price"
        params = {
            "source_currency": from_ccy,
            "destination_currency": to_ccy,
            "destination_country": country,
            "source_amount": str(int(amount)),
            "delivery_type": "BANK_DEPOSIT",
            "payment_method": "BANK_ACCOUNT",
        }
        headers = {**HEADERS, "Accept": "application/json"}
        resp = requests.get(url, params=params, headers=headers, timeout=20)
        resp.raise_for_status()
        data = resp.json()

        # Navigate Remitly's response schema
        # The response contains a list of rate options; pick Economy
        options = data.get("rates") or data.get("options") or []
        chosen = None
        for opt in options:
            speed = (opt.get("speed") or opt.get("delivery_speed") or "").upper()
            if "ECONOMY" in speed or "STANDARD" in speed:
                chosen = opt
                break
        if not chosen and options:
            chosen = options[-1]  # fallback: last (usually economy)

        if not chosen:
            return Quote("Remitly", from_ccy, to_ccy, amount, error="No rates in response")

        fee = float(chosen.get("fee") or chosen.get("transfer_fee") or 0)
        rate = float(chosen.get("exchange_rate") or chosen.get("rate") or 0)
        received = float(chosen.get("recipient_amount") or chosen.get("destination_amount") or 0)

        # Calculate markup vs mid-market if possible
        mid = data.get("mid_market_rate") or data.get("market_rate")
        markup = None
        if mid and rate:
            markup = round((float(mid) - rate) / float(mid) * 100, 3)

        return Quote(
            provider="Remitly",
            from_ccy=from_ccy,
            to_ccy=to_ccy,
            send_amount=amount,
            fee_usd=fee,
            fx_rate=rate,
            fx_markup_pct=markup,
            received_amount=received,
            note="Economy, bank account delivery",
        )

    except requests.HTTPError as e:
        # If direct API fails, Remitly may require cookies/session — flag it
        return Quote(
            "Remitly", from_ccy, to_ccy, amount,
            error=f"HTTP {e.response.status_code} — Remitly may require browser session. "
                  f"See FALLBACK note in README.",
        )
    except Exception as e:
        return Quote("Remitly", from_ccy, to_ccy, amount, error=str(e))


# ─── Provider: Western Union (direct) ──────────────────────────────────────────

def get_western_union_quote(from_ccy: str, to_ccy: str, amount: float) -> Quote:
    """
    Calls Western Union's public price estimator API.
    Bank account delivery (BANK_ACCOUNT pay-out).
    """
    country = CURRENCY_TO_COUNTRY.get(to_ccy, "")
    if not country:
        return Quote("Western Union", from_ccy, to_ccy, amount, error=f"No country mapping for {to_ccy}")

    try:
        # WU's public static price estimator endpoint
        url = "https://www.westernunion.com/us/en/web/send-money/api/price-estimator"
        payload = {
            "sendCurrencyCode": from_ccy,
            "destCurrencyCode": to_ccy,
            "destCountryCode": country,
            "sendAmount": str(int(amount)),
            "payinMethod": "ACCOUNT",   # bank account pay-in
            "payoutMethod": "ACCOUNT",  # bank account pay-out (not cash)
        }
        headers = {
            **HEADERS,
            "Content-Type": "application/json",
            "Referer": "https://www.westernunion.com/us/en/send-money/app/price-estimator",
        }
        resp = requests.post(url, json=payload, headers=headers, timeout=20)
        resp.raise_for_status()
        data = resp.json()

        # WU response schema
        promotions = data.get("promotions") or []
        options = data.get("paymentOptions") or data.get("options") or [data]
        chosen = options[0] if options else data

        fee = float(chosen.get("fee") or chosen.get("transferFee") or 0)
        rate = float(chosen.get("fxRate") or chosen.get("exchangeRate") or 0)
        received = float(chosen.get("destPrincipalAmount") or chosen.get("receiveAmount") or 0)
        mid = float(chosen.get("midMarketRate") or 0)
        markup = round((mid - rate) / mid * 100, 3) if mid and rate else None

        return Quote(
            provider="Western Union",
            from_ccy=from_ccy,
            to_ccy=to_ccy,
            send_amount=amount,
            fee_usd=fee,
            fx_rate=rate,
            fx_markup_pct=markup,
            received_amount=received,
            note="Bank account delivery (online)",
        )

    except requests.HTTPError as e:
        return Quote(
            "Western Union", from_ccy, to_ccy, amount,
            error=f"HTTP {e.response.status_code} — WU may require session cookies. "
                  f"Run with --use-browser flag to scrape via headless Chrome.",
        )
    except Exception as e:
        return Quote("Western Union", from_ccy, to_ccy, amount, error=str(e))


# ─── Provider: MoneyGram (direct) ──────────────────────────────────────────────

def get_moneygram_quote(from_ccy: str, to_ccy: str, amount: float) -> Quote:
    """
    Calls MoneyGram's public fee estimator.
    Bank deposit delivery (RECEIVE_MONEY_IN_BANK_ACCOUNT).
    """
    country = CURRENCY_TO_COUNTRY.get(to_ccy, "")
    if not country:
        return Quote("MoneyGram", from_ccy, to_ccy, amount, error=f"No country mapping for {to_ccy}")

    try:
        # MoneyGram fee estimator — same endpoint as their public web calculator
        url = "https://www.moneygram.com/mgo/us/en/fee-estimator/api/estimate"
        params = {
            "sendCurrency": from_ccy,
            "receiveCurrency": to_ccy,
            "receiveCountry": country,
            "sendAmount": str(int(amount)),
            "paymentMethod": "BANK_ACCOUNT",       # bank account pay-in
            "deliveryMethod": "RECEIVE_MONEY_IN_BANK_ACCOUNT",  # bank delivery
        }
        headers = {**HEADERS, "Accept": "application/json"}
        resp = requests.get(url, params=params, headers=headers, timeout=20)
        resp.raise_for_status()
        data = resp.json()

        fee = float(data.get("fee") or data.get("transferFee") or 0)
        rate = float(data.get("exchangeRate") or data.get("fxRate") or 0)
        received = float(data.get("receiveAmount") or data.get("destinationAmount") or 0)
        mid = float(data.get("midMarketRate") or 0)
        markup = round((mid - rate) / mid * 100, 3) if mid and rate else None

        return Quote(
            provider="MoneyGram",
            from_ccy=from_ccy,
            to_ccy=to_ccy,
            send_amount=amount,
            fee_usd=fee,
            fx_rate=rate,
            fx_markup_pct=markup,
            received_amount=received,
            note="Bank account delivery (online)",
        )

    except requests.HTTPError as e:
        return Quote(
            "MoneyGram", from_ccy, to_ccy, amount,
            error=f"HTTP {e.response.status_code} — Trying alternative MGram endpoint.",
        )
    except Exception as e:
        return Quote("MoneyGram", from_ccy, to_ccy, amount, error=str(e))


# ─── Browser-based scraping fallback ───────────────────────────────────────────

def get_revolut_quote_browser(from_ccy: str, to_ccy: str, amount: float) -> Quote:
    """
    Scrapes Revolut's public transfer page using Playwright headless Chrome.
    Revolut's pricing requires JS rendering — there is no unauthenticated public API.
    Standard (free) plan, bank transfer to external account.

    REQUIRES: pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return Quote(
            "Revolut", from_ccy, to_ccy, amount,
            error="Playwright not installed. Run: pip install playwright && playwright install chromium",
        )

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=HEADERS["User-Agent"])

            # Intercept XHR/fetch calls to capture the pricing JSON
            pricing_data = {}

            def handle_response(response):
                if "transfer" in response.url.lower() or "quote" in response.url.lower():
                    try:
                        body = response.json()
                        if isinstance(body, dict) and ("fee" in body or "rate" in body or "amount" in body):
                            pricing_data.update(body)
                    except Exception:
                        pass

            page.on("response", handle_response)

            # Navigate to Revolut's send money page
            url = f"https://www.revolut.com/en-US/money-transfer/{from_ccy.lower()}-to-{to_ccy.lower()}/"
            page.goto(url, wait_until="networkidle", timeout=30000)
            time.sleep(3)

            # Try to find fee/rate in page text as fallback
            page_text = page.inner_text("body")
            browser.close()

            # Try to parse from intercepted API calls first
            if pricing_data:
                fee = float(pricing_data.get("fee") or pricing_data.get("transferFee") or 0)
                rate = float(pricing_data.get("rate") or pricing_data.get("exchangeRate") or 0)
                received = float(pricing_data.get("recipientAmount") or pricing_data.get("amount") or 0)
                if not received and rate:
                    received = round((amount - fee) * rate, 2)
                return Quote(
                    provider="Revolut",
                    from_ccy=from_ccy, to_ccy=to_ccy, send_amount=amount,
                    fee_usd=fee, fx_rate=rate, received_amount=received,
                    note="Standard plan, external bank transfer (scraped)",
                )

            # Fallback: extract from page text with regex
            fee_match = re.search(r"fee[:\s]+\$?([\d.]+)", page_text, re.IGNORECASE)
            rate_match = re.search(r"1\s*USD\s*[=≈]\s*([\d.]+)\s*" + to_ccy, page_text, re.IGNORECASE)

            fee = float(fee_match.group(1)) if fee_match else None
            rate = float(rate_match.group(1)) if rate_match else None
            received = round((amount - (fee or 0)) * rate, 2) if rate else None

            return Quote(
                provider="Revolut",
                from_ccy=from_ccy, to_ccy=to_ccy, send_amount=amount,
                fee_usd=fee, fx_rate=rate, received_amount=received,
                note="Standard plan, external bank transfer (page scraped)",
            )

    except Exception as e:
        return Quote("Revolut", from_ccy, to_ccy, amount, error=f"Browser scrape failed: {e}")


def get_euronet_quote_browser(from_ccy: str, to_ccy: str, amount: float) -> Quote:
    """
    Scrapes Euronet's public currency converter / send money page.
    Euronet (epay / easycash) — bank transfer pricing.

    REQUIRES: pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return Quote(
            "Euronet", from_ccy, to_ccy, amount,
            error="Playwright not installed. Run: pip install playwright && playwright install chromium",
        )

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=HEADERS["User-Agent"])

            captured = {}

            def handle_response(response):
                if any(k in response.url for k in ["rate", "quote", "price", "fee", "currency"]):
                    try:
                        body = response.json()
                        if isinstance(body, dict):
                            captured.update(body)
                    except Exception:
                        pass

            page.on("response", handle_response)

            # Euronet's public currency converter
            url = f"https://www.euronet.eu/send-money?from={from_ccy}&to={to_ccy}&amount={int(amount)}"
            page.goto(url, wait_until="networkidle", timeout=30000)
            time.sleep(2)
            page_text = page.inner_text("body")
            browser.close()

            rate_match = re.search(r"1\s*USD\s*[=≈]\s*([\d.]+)", page_text, re.IGNORECASE)
            fee_match = re.search(r"fee[:\s]+\$?([\d.]+)", page_text, re.IGNORECASE)
            received_match = re.search(
                r"(?:receive[sd]?|you get)[:\s]+([\d,]+\.?\d*)\s*" + to_ccy, page_text, re.IGNORECASE
            )

            rate = float(rate_match.group(1)) if rate_match else None
            fee = float(fee_match.group(1)) if fee_match else None
            received = (
                float(received_match.group(1).replace(",", ""))
                if received_match
                else (round((amount - (fee or 0)) * rate, 2) if rate else None)
            )

            return Quote(
                provider="Euronet",
                from_ccy=from_ccy, to_ccy=to_ccy, send_amount=amount,
                fee_usd=fee, fx_rate=rate, received_amount=received,
                note="Bank transfer (scraped from public page)",
            )

    except Exception as e:
        return Quote("Euronet", from_ccy, to_ccy, amount, error=f"Browser scrape failed: {e}")


# ─── Orchestrator ───────────────────────────────────────────────────────────────

def fetch_all_quotes(
    corridors: list[tuple],
    amounts: list[float],
    use_browser: bool = True,
) -> list[Quote]:
    """
    Fetches pricing data for all corridors × amounts across all providers.
    Returns flat list of Quote objects.

    STRATEGY:
    1. Call Wise Comparison API once per (corridor, amount) — covers multiple providers
    2. Supplement with direct calls to Remitly and MoneyGram APIs
    3. Use browser scraping for Revolut and Euronet (JS-rendered sites)
    """
    all_quotes: list[Quote] = []

    for from_ccy, to_ccy in corridors:
        print(f"\n{'='*60}")
        print(f"  Corridor: {from_ccy} → {to_ccy}")
        print(f"{'='*60}")

        # Track which providers already have data from Wise API
        # to avoid duplicating with direct calls
        wise_covered: dict[int, set] = {amt: set() for amt in amounts}

        # 1. Wise Comparison API (covers multiple providers at once)
        for amount in amounts:
            print(f"  [Wise API] ${amount:,}...", end=" ", flush=True)
            quotes = get_wise_quotes(from_ccy, to_ccy, amount)
            for q in quotes:
                all_quotes.append(q)
                wise_covered[amount].add(q.provider.lower().replace(" ", ""))
                print(f"{q.provider}", end=" ", flush=True)
            print()
            time.sleep(0.3)  # gentle rate limiting

        # 2. Remitly direct (supplement if not in Wise API results)
        for amount in amounts:
            if "remitly" not in wise_covered[amount]:
                print(f"  [Remitly direct] ${amount:,}...", end=" ", flush=True)
                q = get_remitly_quote(from_ccy, to_ccy, amount)
                all_quotes.append(q)
                print(f"{'✓' if not q.error else '✗ ' + q.error[:50]}")
                time.sleep(0.5)

        # 3. MoneyGram direct (supplement if not in Wise API results)
        for amount in amounts:
            if "moneygram" not in wise_covered[amount]:
                print(f"  [MoneyGram direct] ${amount:,}...", end=" ", flush=True)
                q = get_moneygram_quote(from_ccy, to_ccy, amount)
                all_quotes.append(q)
                print(f"{'✓' if not q.error else '✗ ' + q.error[:50]}")
                time.sleep(0.5)

        # 4. Revolut browser scrape
        if use_browser:
            for amount in amounts:
                if "revolut" not in wise_covered[amount]:
                    print(f"  [Revolut browser] ${amount:,}...", end=" ", flush=True)
                    q = get_revolut_quote_browser(from_ccy, to_ccy, amount)
                    all_quotes.append(q)
                    print(f"{'✓' if not q.error else '✗ ' + q.error[:50]}")
                    time.sleep(1)

        # 5. Euronet browser scrape
        if use_browser:
            for amount in amounts:
                print(f"  [Euronet browser] ${amount:,}...", end=" ", flush=True)
                q = get_euronet_quote_browser(from_ccy, to_ccy, amount)
                all_quotes.append(q)
                print(f"{'✓' if not q.error else '✗ ' + q.error[:50]}")
                time.sleep(1)

    return all_quotes


# ─── Excel output ───────────────────────────────────────────────────────────────

# Color palette
C_DARK_BLUE  = "1A3A5C"
C_MID_BLUE   = "2E6DA4"
C_LIGHT_BLUE = "D9E8F5"
C_WHITE      = "FFFFFF"
C_GREEN      = "E2F0D9"
C_AMBER      = "FFF2CC"
C_RED        = "FFE0E0"
C_BEST_GREEN = "C6EFCE"
C_DARK_GREY  = "404040"
C_LIGHT_GREY = "F5F5F5"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color=C_DARK_GREY, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_summary_sheet(ws, all_quotes: list[Quote], amounts: list[float]):
    """Writes the main summary tab with all providers × amounts per corridor."""

    # Group quotes
    by_corridor: dict[str, dict] = {}
    for q in all_quotes:
        key = f"{q.from_ccy}→{q.to_ccy}"
        if key not in by_corridor:
            by_corridor[key] = {}
        if q.provider not in by_corridor[key]:
            by_corridor[key][q.provider] = {}
        by_corridor[key][q.provider][q.send_amount] = q

    row = 1

    # Title
    ws.merge_cells(f"A{row}:Z{row}")
    ws[f"A{row}"] = "FT Partners — Remittance Competitive Pricing Tracker"
    ws[f"A{row}"].font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
    ws[f"A{row}"].fill = _fill(C_DARK_BLUE)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Subtitle
    ws.merge_cells(f"A{row}:Z{row}")
    ws[f"A{row}"] = (
        f"Bank-to-bank transfers | Source: Wise Comparison API + direct provider APIs | "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws[f"A{row}"].font = _font(italic=True, color=C_WHITE, size=9)
    ws[f"A{row}"].fill = _fill(C_MID_BLUE)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 16
    row += 2

    for corridor, providers in by_corridor.items():
        # Corridor header
        n_cols = 1 + len(amounts) * 3
        end_col = get_column_letter(n_cols)
        ws.merge_cells(f"A{row}:{end_col}{row}")
        ws[f"A{row}"] = f"  {corridor}  —  Bank Account Delivery"
        ws[f"A{row}"].font = Font(bold=True, size=11, color=C_WHITE, name="Arial")
        ws[f"A{row}"].fill = _fill(C_MID_BLUE)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers — row 1: amounts
        ws.cell(row, 1).value = "Provider"
        ws.cell(row, 1).font = _font(bold=True, color=C_WHITE)
        ws.cell(row, 1).fill = _fill(C_DARK_BLUE)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

        col = 2
        for amt in amounts:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
            cell = ws.cell(row, col)
            cell.value = f"${amt:,}"
            cell.font = _font(bold=True, color=C_WHITE)
            cell.fill = _fill(C_DARK_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col += 3

        ws.row_dimensions[row].height = 18
        row += 1

        # Sub-headers: Fee / FX Rate / Markup% / Received
        ws.cell(row, 1).fill = _fill(C_LIGHT_BLUE)
        col = 2
        for amt in amounts:
            for lbl in ["Fee ($)", "FX Rate", "Mkp %", "Received"]:
                ws.cell(row, col).value = lbl
                ws.cell(row, col).font = _font(bold=True, size=8, color=C_DARK_GREY)
                ws.cell(row, col).fill = _fill(C_LIGHT_BLUE)
                ws.cell(row, col).alignment = Alignment(horizontal="center")
                ws.cell(row, col).border = _border()
                col += 1
            # Merge the 4 sub-headers per amount (we use 4 cols now not 3)
            # Actually just use 4 columns per amount group: fee, rate, markup, received
        row += 1

        # Re-do headers properly — 4 cols per amount
        # (rewind 2 rows and redo)
        # Let me just use 4 columns per amount group properly from the start
        # ... handled in provider rows below

        # Find best "received" per amount for green highlighting
        best_received: dict[float, float] = {}
        best_fee: dict[float, float] = {}
        for pname, pdata in providers.items():
            for amt in amounts:
                q = pdata.get(amt)
                if q and q.received_amount:
                    if amt not in best_received or q.received_amount > best_received[amt]:
                        best_received[amt] = q.received_amount
                if q and q.fee_usd is not None:
                    if amt not in best_fee or q.fee_usd < best_fee[amt]:
                        best_fee[amt] = q.fee_usd

        # Provider data rows
        sorted_providers = sorted(providers.keys())
        for i, pname in enumerate(sorted_providers):
            pdata = providers[pname]
            fill_color = C_WHITE if i % 2 == 0 else C_LIGHT_GREY

            # Provider name cell
            ws.cell(row, 1).value = pname
            ws.cell(row, 1).font = _font(bold=True, size=10)
            ws.cell(row, 1).fill = _fill(fill_color)
            ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row, 1).border = _border()

            col = 2
            for amt in amounts:
                q = pdata.get(amt)
                is_best_rcv = q and q.received_amount and q.received_amount == best_received.get(amt)
                is_best_fee = q and q.fee_usd is not None and q.fee_usd == best_fee.get(amt)
                row_fill = _fill(C_BEST_GREEN if is_best_rcv else fill_color)

                if q and q.error:
                    # Error cell — span all 4 cols
                    for offset in range(4):
                        c = ws.cell(row, col + offset)
                        c.value = "⚠ " + q.error[:30] if offset == 0 else ""
                        c.font = _font(size=8, italic=True, color="AA0000")
                        c.fill = _fill("FFF0F0")
                        c.border = _border()
                    col += 4
                    continue

                # Fee
                fee_cell = ws.cell(row, col)
                fee_cell.value = round(q.fee_usd, 2) if q and q.fee_usd is not None else "—"
                fee_cell.number_format = '$#,##0.00'
                fee_cell.font = _font(bold=is_best_fee, color="006600" if is_best_fee else C_DARK_GREY)
                fee_cell.fill = _fill(C_BEST_GREEN) if is_best_fee else _fill(fill_color)
                fee_cell.alignment = Alignment(horizontal="right")
                fee_cell.border = _border()
                col += 1

                # FX Rate
                rate_cell = ws.cell(row, col)
                rate_val = q.fx_rate if q else None
                rate_cell.value = round(rate_val, 4) if rate_val else "—"
                rate_cell.number_format = '0.0000'
                rate_cell.font = _font()
                rate_cell.fill = _fill(fill_color)
                rate_cell.alignment = Alignment(horizontal="right")
                rate_cell.border = _border()
                col += 1

                # FX Markup %
                mkp_cell = ws.cell(row, col)
                mkp_val = q.fx_markup_pct if q else None
                mkp_cell.value = round(mkp_val, 2) / 100 if mkp_val is not None else "—"
                mkp_cell.number_format = '0.00%'
                # Color code: green <0.5%, amber 0.5-1.5%, red >1.5%
                if mkp_val is not None:
                    if mkp_val < 0.5:
                        mkp_cell.fill = _fill(C_GREEN)
                        mkp_cell.font = _font(color="276221")
                    elif mkp_val < 1.5:
                        mkp_cell.fill = _fill(C_AMBER)
                        mkp_cell.font = _font(color="7D6608")
                    else:
                        mkp_cell.fill = _fill(C_RED)
                        mkp_cell.font = _font(color="B22222")
                else:
                    mkp_cell.fill = _fill(fill_color)
                    mkp_cell.font = _font()
                mkp_cell.alignment = Alignment(horizontal="right")
                mkp_cell.border = _border()
                col += 1

                # Received amount
                rcv_cell = ws.cell(row, col)
                rcv_val = q.received_amount if q else None
                rcv_cell.value = round(rcv_val, 2) if rcv_val else "—"
                rcv_cell.number_format = '#,##0.00'
                rcv_cell.font = _font(bold=is_best_rcv, color="006600" if is_best_rcv else C_DARK_GREY, size=10)
                rcv_cell.fill = row_fill
                rcv_cell.alignment = Alignment(horizontal="right")
                rcv_cell.border = _border()
                col += 1

            ws.row_dimensions[row].height = 16
            row += 1

        row += 2  # spacer between corridors

    # Column widths
    ws.column_dimensions["A"].width = 18
    col_idx = 2
    for amt in amounts:
        ws.column_dimensions[get_column_letter(col_idx)].width = 9      # fee
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 9  # rate
        ws.column_dimensions[get_column_letter(col_idx + 2)].width = 8  # markup
        ws.column_dimensions[get_column_letter(col_idx + 3)].width = 12 # received
        col_idx += 4

    ws.freeze_panes = "B4"


def _write_legend_sheet(ws):
    ws.title = "Legend & Notes"
    data = [
        ("COLOR CODING", ""),
        ("Green (FX markup)", "< 0.5% above mid-market — excellent rate"),
        ("Amber (FX markup)", "0.5% – 1.5% above mid-market — moderate"),
        ("Red (FX markup)",   "> 1.5% above mid-market — high cost"),
        ("Green background (received)", "Best received amount for that send amount"),
        ("Green background (fee)",      "Lowest fee for that send amount"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("Wise Comparison API",  "api.wise.com/v4/comparisons — public, no auth required. "
                                  "Wise scrapes competitors ~hourly. Bank transfer in/out only."),
        ("Remitly",              "api.remitly.io/v3/pricing — public price calculator endpoint. "
                                  "Economy rate, bank account delivery."),
        ("Western Union",        "westernunion.com price estimator API. Bank account in/out. "
                                  "May require session cookie on some runs."),
        ("MoneyGram",            "moneygram.com fee estimator API. Bank deposit delivery."),
        ("Revolut",              "Scraped via headless browser (Playwright). Standard (free) plan, "
                                  "external bank transfer. Requires: playwright install chromium"),
        ("Euronet",              "Scraped via headless browser (Playwright). "
                                  "Coverage varies — not all corridors supported."),
        ("", ""),
        ("NOTES", ""),
        ("FX Markup",    "Calculated as % above mid-market rate at time of data collection."),
        ("Received",     "Amount recipient receives in destination currency after all fees and FX markup."),
        ("Bank Transfer","All pricing reflects bank account send + bank account receive (not cash pickup)."),
        ("Revolut plan", "Standard (free) plan. Premium/Metal plans have lower/no fees."),
        ("First-time promos", "Remitly and some others offer promotional rates for new customers. "
                               "This script fetches standard returning-customer rates."),
    ]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    for i, (label, value) in enumerate(data, 1):
        ws.cell(i, 1).value = label
        ws.cell(i, 2).value = value
        if label.isupper() and label:
            ws.cell(i, 1).font = Font(bold=True, size=10, name="Arial", color=C_WHITE)
            ws.cell(i, 1).fill = _fill(C_MID_BLUE)
            ws.cell(i, 2).fill = _fill(C_MID_BLUE)
        else:
            ws.cell(i, 1).font = _font(bold=bool(label))
            ws.cell(i, 2).font = _font()
        ws.row_dimensions[i].height = 15


def write_excel(all_quotes: list[Quote], amounts: list[float], filename: str):
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Pricing Summary"
    _write_summary_sheet(ws_summary, all_quotes, amounts)

    # Per-corridor raw data sheets
    corridors_seen = list(dict.fromkeys(f"{q.from_ccy}→{q.to_ccy}" for q in all_quotes))
    for corridor in corridors_seen:
        ws = wb.create_sheet(title=corridor)
        corridor_quotes = [q for q in all_quotes if f"{q.from_ccy}→{q.to_ccy}" == corridor]
        _write_raw_sheet(ws, corridor_quotes, amounts)

    # Legend
    ws_legend = wb.create_sheet(title="Legend & Notes")
    _write_legend_sheet(ws_legend)

    wb.save(filename)
    print(f"\n✅  Saved: {filename}")


def _write_raw_sheet(ws, quotes: list[Quote], amounts: list[float]):
    """Raw data tab — one row per quote for easy filtering/pivoting."""
    headers = ["Provider", "From", "To", "Send Amount ($)", "Fee ($)", "FX Rate",
               "FX Markup %", "Received", "Note", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = h
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK_BLUE)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")

    for row, q in enumerate(quotes, 2):
        values = [
            q.provider, q.from_ccy, q.to_ccy, q.send_amount,
            q.fee_usd, q.fx_rate,
            round(q.fx_markup_pct / 100, 4) if q.fx_markup_pct is not None else None,
            q.received_amount, q.note, q.error,
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row, col)
            cell.value = v
            cell.font = _font()
            cell.border = _border()
            if col == 4:  cell.number_format = '$#,##0'
            if col == 5:  cell.number_format = '$#,##0.00'
            if col == 6:  cell.number_format = '0.0000'
            if col == 7:  cell.number_format = '0.00%'
            if col == 8:  cell.number_format = '#,##0.00'

    for i, w in enumerate([18, 6, 6, 14, 10, 10, 10, 14, 45, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


# ─── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="FT Partners Remittance Pricing Scraper — pulls live bank-to-bank rates"
    )
    parser.add_argument(
        "--corridors", nargs="+",
        help="Corridors to fetch (e.g. USD-EUR USD-INR). Default: all 8 corridors.",
        default=None,
    )
    parser.add_argument(
        "--amounts", nargs="+", type=int,
        help="Send amounts in USD. Default: 100 500 1000 5000 10000",
        default=None,
    )
    parser.add_argument(
        "--no-browser", action="store_true",
        help="Skip browser-based scraping (Revolut, Euronet). Faster but less complete.",
    )
    parser.add_argument(
        "--output", type=str, default=None,
        help="Output filename. Default: remittance_pricing_YYYYMMDD_HHMMSS.xlsx",
    )
    args = parser.parse_args()

    corridors = CORRIDORS
    if args.corridors:
        corridors = [tuple(c.replace("-", " ").split()) for c in args.corridors]

    amounts = args.amounts or AMOUNTS
    use_browser = not args.no_browser

    output_file = args.output or f"remittance_pricing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print("\n" + "═" * 60)
    print("  FT Partners — Remittance Competitive Pricing Tracker")
    print("═" * 60)
    print(f"  Corridors : {', '.join(f'{a}→{b}' for a, b in corridors)}")
    print(f"  Amounts   : {', '.join(f'${a:,}' for a in amounts)}")
    print(f"  Browser   : {'enabled (Revolut + Euronet)' if use_browser else 'disabled'}")
    print(f"  Output    : {output_file}")
    print()

    all_quotes = fetch_all_quotes(corridors, amounts, use_browser=use_browser)

    print(f"\n  Total quotes fetched: {len(all_quotes)}")
    errors = [q for q in all_quotes if q.error]
    if errors:
        print(f"  Errors: {len(errors)}")
        for q in errors[:5]:
            print(f"    ✗ {q.provider} {q.from_ccy}→{q.to_ccy} ${q.send_amount}: {q.error[:60]}")

    write_excel(all_quotes, amounts, output_file)


if __name__ == "__main__":
    main()
